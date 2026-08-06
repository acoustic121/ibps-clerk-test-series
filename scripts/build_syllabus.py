import openpyxl
import json
import os

BASE = "/Users/aman.singh/Documents/banking"
OUT = "/Users/aman.singh/Documents/banking/test-series-app/server/data/syllabus.json"


def load_subjects(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    subjects = []
    for name in wb.sheetnames:
        if name == "Overview":
            continue
        ws = wb[name]
        chapters = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            chapters.append(row[1])
        subjects.append({"name": name, "chapters": chapters})
    return subjects


def main():
    syllabus = {
        "prelims": {
            "label": "IBPS Clerk Prelims",
            "totalMarks": 100,
            "durationMinutes": 60,
            "subjects": load_subjects(os.path.join(BASE, "IBPS_Clerk_Prelims_Syllabus.xlsx")),
        },
        "mains": {
            "label": "IBPS Clerk Mains",
            "totalMarks": 200,
            "durationMinutes": 125,
            "subjects": load_subjects(os.path.join(BASE, "IBPS_Clerk_Mains_Syllabus.xlsx")),
        },
        "real_prelims": {
            "label": "Real Prelims (Previous Year Papers)",
            "totalMarks": 100,
            "durationMinutes": 60,
            "subjects": [],
        },
        "real_mains": {
            "label": "Real Mains (Previous Year Papers)",
            "totalMarks": 200,
            "durationMinutes": 125,
            "subjects": [],
        },
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(syllabus, f, indent=2, ensure_ascii=False)
    print(f"Wrote syllabus to {OUT}")
    for track, data in syllabus.items():
        print(f"  {track}: {sum(len(s['chapters']) for s in data['subjects'])} chapters across {len(data['subjects'])} subjects")


if __name__ == "__main__":
    main()
